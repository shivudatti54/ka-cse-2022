#!/usr/bin/env python3
"""Generate VTU content review and syllabus coverage Excel workbooks.

Inputs:
- VTU syllabus PDFs (Sem 3-4, 5, 6, 7)
- Content pack structure under vtu-2022-scheme/cse

Outputs:
- review_exports/vtu_2022_theory_content_review_tracker.xlsx
- review_exports/vtu_2022_theory_syllabus_coverage.xlsx
"""

from __future__ import annotations

import json
import re
import statistics
import subprocess
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rapidfuzz import fuzz


CONTENT_ROOT = Path(
    "/Users/ggoudar/Documents/GitHub/smalltools/clawdBot/study-app-template/content-packs/vtu-2022-scheme/cse"
)
SYLLABUS_PDFS = {
    "sem34": Path(
        "/Users/ggoudar/Documents/GitHub/smalltools/vtu_engg_model_answers/vtu_syllabus/2022_Scheme/Computer_Science_Engineering/Sem_3-4_Syllabus.pdf"
    ),
    "sem5": Path(
        "/Users/ggoudar/Documents/GitHub/smalltools/vtu_engg_model_answers/vtu_syllabus/2022_Scheme/Computer_Science_Engineering/Sem_5_Syllabus.pdf"
    ),
    "sem6": Path(
        "/Users/ggoudar/Documents/GitHub/smalltools/vtu_engg_model_answers/vtu_syllabus/2022_Scheme/Computer_Science_Engineering/Sem_6_Syllabus.pdf"
    ),
    "sem7": Path(
        "/Users/ggoudar/Documents/GitHub/smalltools/vtu_engg_model_answers/vtu_syllabus/2022_Scheme/Computer_Science_Engineering/Sem_7_Syllabus.pdf"
    ),
}

OUTPUT_DIR = CONTENT_ROOT / "review_exports"
TEXT_CACHE_DIR = OUTPUT_DIR / "_syllabus_text_cache"

MODULE_HEADER_RE = re.compile(
    r"^\s*(?:MODULE|Module)\s*[-–]?\s*(\d)(?:\s*and\s*(\d))?\s*:?[ \t]*(.*)$"
)
CODE_RE = re.compile(r"\b[A-Z]{3,4}\d{3}[A-Z]?\b")


STOP_TOPIC_PHRASES = {
    "rbt",
    "textbook",
    "text book",
    "chapter",
    "hours",
    "module",
    "pedagogy",
    "course objectives",
    "course objective",
}

MATCH_COVERED_THRESHOLD = 75
MATCH_PARTIAL_THRESHOLD = 55


@dataclass
class AppTopic:
    semester: int
    subject_code: str
    subject_name: str
    module_no: int
    module_title: str
    topic_id: str
    topic_title: str
    topic_path: str
    read_words: int
    mcq_count: int
    question_count: int
    flashcard_count: int
    has_visual: bool
    has_summary: bool
    has_full: bool
    has_purpose: bool
    has_memory: bool
    rating: float


@dataclass
class AppSubject:
    semester: int
    code: str
    name: str
    path: Path
    modules: Dict[int, Dict[str, object]] = field(default_factory=dict)
    topics: List[AppTopic] = field(default_factory=list)


@dataclass
class SyllabusModule:
    module_no: int
    module_title: str
    raw_text: str
    topics: List[str]


@dataclass
class SyllabusSubject:
    semester: Optional[int]
    codes: List[str]
    display_title: str
    source: str
    exam_type: str
    practical_only: bool
    modules: Dict[int, SyllabusModule]


def slugify_text(text: str) -> str:
    text = text.lower().strip()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_code(code: str) -> str:
    return re.sub(r"\s+", "", (code or "").upper())


def parse_roman_or_int(value: str) -> Optional[int]:
    value = (value or "").strip().upper()
    if not value:
        return None
    if value.isdigit():
        return int(value)
    roman_map = {"I": 1, "V": 5, "X": 10}
    total = 0
    prev = 0
    for ch in reversed(value):
        if ch not in roman_map:
            return None
        cur = roman_map[ch]
        if cur < prev:
            total -= cur
        else:
            total += cur
            prev = cur
    return total if total > 0 else None


def first_nonempty(lines: List[str]) -> str:
    for line in lines:
        if line.strip():
            return line.strip()
    return ""


def extract_pdf_text(pdf_path: Path, key: str) -> str:
    TEXT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    out_path = TEXT_CACHE_DIR / f"{key}.txt"
    if (not out_path.exists()) or out_path.stat().st_mtime < pdf_path.stat().st_mtime:
        subprocess.run(
            ["pdftotext", "-layout", "-nopgbrk", str(pdf_path), str(out_path)],
            check=True,
        )
    return out_path.read_text(encoding="utf-8", errors="ignore")


def parse_exam_type(header_lines: List[str]) -> str:
    for i, line in enumerate(header_lines):
        if re.search(r"\bexamination\b", line, flags=re.IGNORECASE):
            combined = line
            if i + 1 < len(header_lines):
                combined += " " + header_lines[i + 1]
            if re.search(r"theory|practical", combined, flags=re.IGNORECASE):
                return re.sub(r"\s+", " ", combined).strip()
    combined = " ".join(header_lines)
    if re.search(r"theory|practical", combined, flags=re.IGNORECASE):
        return combined
    return ""


def is_probably_lab_code(code: str) -> bool:
    code = normalize_code(code)
    # Examples: BCSL305, BAIL657C, BCSL606
    return bool(re.match(r"^[A-Z]{2,4}L\d", code))


def clean_module_title(title: str, module_no: int) -> str:
    title = re.sub(r"(?i)\bno\.?\s*of\s*hours\b.*", "", title)
    title = re.sub(r"(?i)\b\d+\s*hours?\b.*", "", title)
    title = re.sub(r"\s+", " ", title).strip(" :-\t")
    if not title:
        return f"Module {module_no}"
    return title


def clean_theory_line(line: str) -> str:
    line = line.strip()
    line = re.sub(r"@#\s*\d+", " ", line)
    line = re.sub(r"\bAnnexure\-?[IVX]+\b", " ", line, flags=re.IGNORECASE)
    line = re.sub(r"\s+", " ", line)
    return line.strip()


def split_topic_phrase(phrase: str) -> List[str]:
    phrase = phrase.strip(" -:\t")
    phrase = re.sub(r"\s+", " ", phrase).strip()
    if not phrase:
        return []
    words = phrase.split()
    if len(words) <= 16:
        return [phrase]
    # If very long, split by " and " as fallback.
    parts = [p.strip() for p in re.split(r"\band\b", phrase, flags=re.IGNORECASE) if p.strip()]
    return parts if len(parts) > 1 else [phrase]


def extract_topics_from_module_text(module_text: str) -> List[str]:
    text = module_text
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"(?i)\btext\s*book\s*\d*\s*:\s*[^.]*", " ", text)
    text = re.sub(r"(?i)\btextbook\s*\d*\s*:\s*[^.]*", " ", text)
    text = re.sub(r"(?i)\bchapter\s*\d+[^.]*", " ", text)
    text = re.sub(r"(?i)\brbt\s*:\s*[^.]*", " ", text)
    text = re.sub(r"\([^)]*hours?[^)]*\)", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\b\d+\s*hours?\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()

    raw_parts = re.split(r"[,:;.]", text)
    topics: List[str] = []
    seen = set()
    for part in raw_parts:
        for candidate in split_topic_phrase(part):
            c = re.sub(r"\s+", " ", candidate).strip(" -\t")
            if not c:
                continue
            lower = c.lower()
            if any(sp in lower for sp in STOP_TOPIC_PHRASES):
                continue
            if re.fullmatch(r"[\d\W_]+", c):
                continue
            if len(c) < 3:
                continue
            if len(re.findall(r"[A-Za-z]", c)) == 0:
                continue
            if lower not in seen:
                seen.add(lower)
                topics.append(c)
    return topics


def parse_syllabus_subjects(text: str, source_key: str) -> List[SyllabusSubject]:
    lines = text.splitlines()
    starts: List[int] = []
    for i, line in enumerate(lines):
        if re.search(r"\b(?:Course|Subject)\s+Code\b", line, flags=re.IGNORECASE):
            starts.append(i)

    subjects: List[SyllabusSubject] = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] if idx + 1 < len(starts) else len(lines)
        section_lines = lines[start:end]
        header_lines = section_lines[:12]
        header_text = " ".join(header_lines)

        # Primary codes are on the "Course Code" line; sometimes one extra equivalent code
        # appears on the next line (e.g., BCS515A / BAI515A).
        codes = CODE_RE.findall(section_lines[0]) if section_lines else []
        for offset in (1, 2):
            if offset >= len(section_lines):
                continue
            line = section_lines[offset]
            if re.search(
                r"\b(Teaching|CIE|SEE|Total|Credits|Examination|Semester|Course objectives|Course objective)\b",
                line,
                flags=re.IGNORECASE,
            ):
                continue
            extra = CODE_RE.findall(line)
            if extra:
                codes.extend(extra)
        codes = list(dict.fromkeys(codes))

        if not codes:
            continue

        norm_codes = [normalize_code(c) for c in codes]

        # Infer subject title from nearby line containing "Semester".
        display_title = ""
        semester: Optional[int] = None

        for j in range(start - 1, max(-1, start - 12), -1):
            line = lines[j].strip()
            if not line:
                continue
            if re.search(r"\bsemester\b", line, flags=re.IGNORECASE):
                display_title = re.sub(r"(?i)\bsemester\b.*$", "", line).strip(" :-\t")
                sem_m = re.search(r"(?i)\bsemester\b\s*[:\-]?\s*([IVX]+|\d+)", line)
                if sem_m:
                    semester = parse_roman_or_int(sem_m.group(1))
                break

        if semester is None:
            sem_m = re.search(r"(?i)\bsemester\b\s*[:\-]?\s*([IVX]+|\d+)", header_text)
            if sem_m:
                semester = parse_roman_or_int(sem_m.group(1))

        if not display_title:
            display_title = first_nonempty([clean_theory_line(x) for x in lines[max(0, start - 6):start]])

        exam_type = parse_exam_type(section_lines[:20])
        exam_l = exam_type.lower()
        practical_only = ("practical" in exam_l) and ("theory" not in exam_l)
        # Fallback for clear lab-coded subjects.
        if all(is_probably_lab_code(c) for c in norm_codes):
            practical_only = True

        # Isolate theory module block by stopping at common section boundaries.
        stop_keywords = [
            "practical component",
            "laboratory outcomes",
            "assessment details",
            "continuous internal evaluation",
            "course outcomes",
            "course outcome",
            "course outcome (course skill set)",
        ]
        stop_at = len(section_lines)
        for i, line in enumerate(section_lines):
            low = line.lower()
            if any(k in low for k in stop_keywords):
                stop_at = i
                break

        theory_lines = section_lines[:stop_at]

        # Parse module blocks.
        module_starts: List[Tuple[int, int, Optional[int], str]] = []
        for i, line in enumerate(theory_lines):
            m = MODULE_HEADER_RE.match(line)
            if not m:
                continue
            m1 = int(m.group(1))
            m2 = int(m.group(2)) if m.group(2) else None
            title_tail = clean_module_title(m.group(3) or "", m1)
            module_starts.append((i, m1, m2, title_tail))

        modules: Dict[int, SyllabusModule] = {}
        for i, (line_idx, m1, m2, title_tail) in enumerate(module_starts):
            body_start = line_idx + 1
            body_end = module_starts[i + 1][0] if i + 1 < len(module_starts) else len(theory_lines)
            body_lines = [clean_theory_line(x) for x in theory_lines[body_start:body_end]]
            body_lines = [x for x in body_lines if x]
            body_text = " ".join(body_lines)
            topics = extract_topics_from_module_text(body_text)

            module_title = title_tail if title_tail and title_tail.lower() != f"module {m1}" else ""
            if not module_title and topics:
                module_title = topics[0] if len(topics[0].split()) <= 8 else f"Module {m1}"
            if not module_title:
                module_title = f"Module {m1}"

            modules[m1] = SyllabusModule(
                module_no=m1,
                module_title=module_title,
                raw_text=body_text,
                topics=topics,
            )

            # Rare case: "Module 4 and 5" in theory section.
            if m2 is not None:
                modules[m2] = SyllabusModule(
                    module_no=m2,
                    module_title=f"Module {m2}",
                    raw_text=body_text,
                    topics=list(topics),
                )

        subject = SyllabusSubject(
            semester=semester,
            codes=norm_codes,
            display_title=display_title,
            source=source_key,
            exam_type=exam_type,
            practical_only=practical_only,
            modules=modules,
        )
        subjects.append(subject)

    return subjects


def read_json(path: Path) -> object:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def count_words(path: Path) -> int:
    if not path.exists():
        return 0
    text = path.read_text(encoding="utf-8", errors="ignore")
    return len(re.findall(r"\b\w+\b", text))


def count_json_items(path: Path) -> int:
    if not path.exists():
        return 0
    try:
        data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
    except Exception:
        return 0

    if isinstance(data, list):
        return len(data)
    if isinstance(data, dict):
        for key in ("questions", "mcqs", "flashcards", "items", "data"):
            value = data.get(key)
            if isinstance(value, list):
                return len(value)
        return len(data)
    return 0


def topic_quality_rating(topic_dir: Path) -> Tuple[float, Dict[str, object]]:
    read_md = topic_dir / "read.md"
    summary_md = topic_dir / "summary.md"
    full_md = topic_dir / "full.md"
    purpose_md = topic_dir / "purpose.md"
    memory_json = topic_dir / "memory.json"
    visual_json = topic_dir / "visual.json"
    mcq_json = topic_dir / "mcqs.json"
    questions_json = topic_dir / "questions.json"
    flashcards_json = topic_dir / "flashcards.json"

    read_words = count_words(read_md)
    has_summary = summary_md.exists()
    has_full = full_md.exists()
    has_purpose = purpose_md.exists()
    has_memory = memory_json.exists()
    has_visual = visual_json.exists() or any((topic_dir / "assets").glob("*.svg"))
    mcq_count = count_json_items(mcq_json)
    question_count = count_json_items(questions_json)
    flashcard_count = count_json_items(flashcards_json)

    score = 0.0
    if read_words >= 180:
        score += 1.8
    elif read_words > 0:
        score += 1.0

    if has_summary:
        score += 0.6
    if has_full:
        score += 0.5
    if has_purpose:
        score += 0.3
    if has_memory:
        score += 0.2
    if has_visual:
        score += 0.5
    if mcq_count > 0:
        score += 0.8
    if question_count > 0:
        score += 0.8
    if flashcard_count > 0:
        score += 0.5

    rating = round(min(5.0, score), 1)

    return rating, {
        "read_words": read_words,
        "mcq_count": mcq_count,
        "question_count": question_count,
        "flashcard_count": flashcard_count,
        "has_visual": has_visual,
        "has_summary": has_summary,
        "has_full": has_full,
        "has_purpose": has_purpose,
        "has_memory": has_memory,
    }


def load_app_subjects() -> Dict[str, AppSubject]:
    app_subjects: Dict[str, AppSubject] = {}

    for meta_path in sorted(CONTENT_ROOT.glob("sem-*/*/_meta.json")):
        subject_dir = meta_path.parent
        meta = read_json(meta_path)
        if not isinstance(meta, dict):
            continue

        code = normalize_code(str(meta.get("code", "")))
        if not code:
            continue
        semester = int(meta.get("semester", 0) or 0)
        name = re.sub(r"\s+", " ", str(meta.get("name", "")).strip())

        subject = AppSubject(
            semester=semester,
            code=code,
            name=name,
            path=subject_dir,
            modules={},
            topics=[],
        )

        chapters_index_path = subject_dir / "chapters" / "_index.json"
        chapters_index = read_json(chapters_index_path)
        chapters = chapters_index.get("chapters", []) if isinstance(chapters_index, dict) else []

        for chapter in chapters:
            if not isinstance(chapter, dict):
                continue
            chapter_id = str(chapter.get("id", ""))
            module_match = re.search(r"module-(\d+)", chapter_id)
            if not module_match:
                continue
            module_no = int(module_match.group(1))
            module_title = str(chapter.get("title", f"Module {module_no}"))

            module_index_path = subject_dir / "chapters" / chapter_id / "_index.json"
            module_index = read_json(module_index_path)
            topics = module_index.get("topics", []) if isinstance(module_index, dict) else []

            module_bucket = subject.modules.setdefault(
                module_no,
                {
                    "module_no": module_no,
                    "module_title": module_title,
                    "topics": [],
                },
            )

            for topic in topics:
                if not isinstance(topic, dict):
                    continue
                topic_id = str(topic.get("id", ""))
                topic_title = str(topic.get("title", topic_id)).strip()
                topic_dir = subject_dir / "chapters" / chapter_id / "topics" / topic_id

                rating, metrics = topic_quality_rating(topic_dir)
                topic_rel_path = str(topic_dir.relative_to(CONTENT_ROOT.parent))

                app_topic = AppTopic(
                    semester=semester,
                    subject_code=code,
                    subject_name=name,
                    module_no=module_no,
                    module_title=module_title,
                    topic_id=topic_id,
                    topic_title=topic_title,
                    topic_path=topic_rel_path,
                    read_words=int(metrics["read_words"]),
                    mcq_count=int(metrics["mcq_count"]),
                    question_count=int(metrics["question_count"]),
                    flashcard_count=int(metrics["flashcard_count"]),
                    has_visual=bool(metrics["has_visual"]),
                    has_summary=bool(metrics["has_summary"]),
                    has_full=bool(metrics["has_full"]),
                    has_purpose=bool(metrics["has_purpose"]),
                    has_memory=bool(metrics["has_memory"]),
                    rating=rating,
                )

                module_bucket["topics"].append(app_topic)
                subject.topics.append(app_topic)

        app_subjects[code] = subject

    return app_subjects


def coverage_status(score: float) -> str:
    if score >= MATCH_COVERED_THRESHOLD:
        return "Covered"
    if score >= MATCH_PARTIAL_THRESHOLD:
        return "Partial"
    return "Not Covered"


def best_match_for_phrase(phrase: str, candidate_topics: List[AppTopic]) -> Tuple[Optional[AppTopic], float]:
    if not candidate_topics:
        return None, 0.0
    norm_phrase = slugify_text(phrase)
    best_topic = None
    best_score = -1.0

    for topic in candidate_topics:
        norm_topic = slugify_text(topic.topic_title)
        score = max(
            fuzz.token_set_ratio(norm_phrase, norm_topic),
            fuzz.partial_ratio(norm_phrase, norm_topic),
        )
        if score > best_score:
            best_score = float(score)
            best_topic = topic

    return best_topic, round(best_score, 1)


def build_syllabus_maps(subjects: List[SyllabusSubject]) -> Dict[str, SyllabusSubject]:
    by_code: Dict[str, SyllabusSubject] = {}

    for subj in subjects:
        for code in subj.codes:
            code = normalize_code(code)
            if not code:
                continue
            existing = by_code.get(code)
            if existing is None:
                by_code[code] = subj
                continue
            # Prefer subject with more module data.
            if len(subj.modules) > len(existing.modules):
                by_code[code] = subj

    return by_code


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, max_width: int = 60) -> None:
    widths: Dict[int, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for c in row:
            if c.value is None:
                continue
            value = str(c.value)
            widths[c.column] = max(widths.get(c.column, 0), min(len(value) + 2, max_width))
    for col_idx, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width


def generate_workbooks(
    app_subjects: Dict[str, AppSubject],
    syllabus_by_code: Dict[str, SyllabusSubject],
) -> Tuple[Path, Path, Dict[str, Dict[str, object]], List[Dict[str, object]]]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    review_path = OUTPUT_DIR / "vtu_2022_theory_content_review_tracker.xlsx"
    coverage_path = OUTPUT_DIR / "vtu_2022_theory_syllabus_coverage.xlsx"

    # Build reverse lookup for syllabus phrases by subject+module.
    syllabus_topics_by_subject_module: Dict[Tuple[str, int], List[str]] = defaultdict(list)
    for code, subj in syllabus_by_code.items():
        if subj.practical_only:
            continue
        for m_no, module in subj.modules.items():
            syllabus_topics_by_subject_module[(code, m_no)].extend(module.topics)

    # Detailed review workbook
    wb_review = Workbook()
    default_sheet = wb_review.active
    wb_review.remove(default_sheet)

    ws_intro = wb_review.create_sheet("README")
    ws_intro.append(["Workbook", "Description"])
    ws_intro.append(["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_intro.append([
        "Purpose",
        "Topic-wise QA review tracker for VTU 2022 theory content with tester sign-off columns.",
    ])
    ws_intro.append([
        "Auto Rating (1-5)",
        "Calculated from availability of read/summary/full notes, questions, MCQs, flashcards and visuals.",
    ])
    ws_intro.append([
        "Auto Coverage Tag",
        "Fuzzy-matched against syllabus module topics (Covered/Partial/Not Covered).",
    ])
    style_header(ws_intro, 1)
    autosize_columns(ws_intro)

    ws_all = wb_review.create_sheet("All Subjects")
    ws_all.append(
        [
            "Semester",
            "Subject Code",
            "Subject Name",
            "Module",
            "Module Title",
            "Topic Title",
            "Topic Path",
            "Read Words",
            "MCQs",
            "Questions",
            "Flashcards",
            "Has Visual",
            "Auto Rating (1-5)",
            "Syllabus Topic Match",
            "Match Score",
            "Auto Coverage Tag",
            "Tester Status",
            "Tester Name",
            "Tester Sign-off Date",
            "Reviewer Sign-off",
            "Comments",
        ]
    )
    style_header(ws_all, 1)

    subject_summary: Dict[str, Dict[str, object]] = {}

    for code in sorted(app_subjects.keys(), key=lambda c: (app_subjects[c].semester, c)):
        app_subj = app_subjects[code]
        sheet_name = code if code not in wb_review.sheetnames else f"{code}_{app_subj.semester}"
        ws = wb_review.create_sheet(sheet_name[:31])
        ws.append(
            [
                "Semester",
                "Subject Code",
                "Subject Name",
                "Module",
                "Module Title",
                "Topic Title",
                "Topic Path",
                "Read Words",
                "MCQs",
                "Questions",
                "Flashcards",
                "Has Visual",
                "Auto Rating (1-5)",
                "Syllabus Topic Match",
                "Match Score",
                "Auto Coverage Tag",
                "Tester Status",
                "Tester Name",
                "Tester Sign-off Date",
                "Reviewer Sign-off",
                "Comments",
            ]
        )
        style_header(ws, 1)

        coverage_counts = {"Covered": 0, "Partial": 0, "Not Covered": 0}
        ratings = [t.rating for t in app_subj.topics]

        for topic in sorted(app_subj.topics, key=lambda t: (t.module_no, t.topic_title.lower())):
            candidates = syllabus_topics_by_subject_module.get((code, topic.module_no), [])
            # Fallback: all subject syllabus phrases if module-level list unavailable.
            if not candidates and code in syllabus_by_code:
                for mod in syllabus_by_code[code].modules.values():
                    candidates.extend(mod.topics)

            best_phrase = ""
            best_score = 0.0
            if candidates:
                norm_topic = slugify_text(topic.topic_title)
                for phrase in candidates:
                    norm_phrase = slugify_text(phrase)
                    score = max(
                        fuzz.token_set_ratio(norm_topic, norm_phrase),
                        fuzz.partial_ratio(norm_topic, norm_phrase),
                    )
                    if score > best_score:
                        best_score = float(score)
                        best_phrase = phrase

            auto_tag = coverage_status(best_score)
            coverage_counts[auto_tag] += 1

            row = [
                topic.semester,
                topic.subject_code,
                topic.subject_name,
                topic.module_no,
                topic.module_title,
                topic.topic_title,
                topic.topic_path,
                topic.read_words,
                topic.mcq_count,
                topic.question_count,
                topic.flashcard_count,
                "Yes" if topic.has_visual else "No",
                topic.rating,
                best_phrase,
                round(best_score, 1),
                auto_tag,
                "Pending",
                "",
                "",
                "",
                "",
            ]
            ws.append(row)
            ws_all.append(row)

        for target_ws in (ws,):
            for row_num in range(2, target_ws.max_row + 1):
                status = str(target_ws.cell(row=row_num, column=16).value)
                if status == "Covered":
                    fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
                elif status == "Partial":
                    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                else:
                    fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                target_ws.cell(row=row_num, column=16).fill = fill

        autosize_columns(ws)

        subject_summary[code] = {
            "semester": app_subj.semester,
            "subject_name": app_subj.name,
            "topic_count": len(app_subj.topics),
            "avg_rating": round(statistics.mean(ratings), 2) if ratings else 0,
            "covered_topics": coverage_counts["Covered"],
            "partial_topics": coverage_counts["Partial"],
            "not_covered_topics": coverage_counts["Not Covered"],
        }

    for row_num in range(2, ws_all.max_row + 1):
        status = str(ws_all.cell(row=row_num, column=16).value)
        if status == "Covered":
            fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        elif status == "Partial":
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        else:
            fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        ws_all.cell(row=row_num, column=16).fill = fill

    autosize_columns(ws_all)
    wb_review.save(review_path)

    # Coverage workbook
    wb_cov = Workbook()
    ws_sum = wb_cov.active
    ws_sum.title = "Coverage Summary"
    ws_sum.append(
        [
            "Semester",
            "Subject Code",
            "Subject Name",
            "In App Content",
            "Total Syllabus Topics",
            "Covered",
            "Partial",
            "Not Covered",
            "Coverage %",
            "Fully Covered",
            "Avg Content Rating",
            "Notes",
        ]
    )
    style_header(ws_sum, 1)

    ws_topic = wb_cov.create_sheet("Topic Coverage")
    ws_topic.append(
        [
            "Semester",
            "Subject Code",
            "Subject Name",
            "Module",
            "Module Title",
            "Syllabus Topic",
            "Matched App Topic",
            "Match Score",
            "Coverage Status",
            "Matched Topic Path",
            "Matched Topic Rating",
            "Tester Note",
        ]
    )
    style_header(ws_topic, 1)

    ws_missing = wb_cov.create_sheet("Missing Theory Subjects")
    ws_missing.append(
        [
            "Semester",
            "Subject Code",
            "Syllabus Subject Title",
            "Exam Type",
            "Reason",
        ]
    )
    style_header(ws_missing, 1)

    # Build theory syllabus codes list (excluding practical-only and labs)
    theory_syllabus_codes: Dict[str, SyllabusSubject] = {}
    for code, subj in syllabus_by_code.items():
        if subj.practical_only:
            continue
        if is_probably_lab_code(code):
            continue
        theory_syllabus_codes[code] = subj

    detailed_topic_rows: List[Dict[str, object]] = []

    for code in sorted(theory_syllabus_codes.keys(), key=lambda c: ((theory_syllabus_codes[c].semester or 0), c)):
        subj = theory_syllabus_codes[code]
        app_subj = app_subjects.get(code)

        total = covered = partial = not_covered = 0
        notes = ""
        ratings: List[float] = []

        if app_subj is None:
            notes = "Theory subject not found in app content"
            for module_no in sorted(subj.modules.keys()):
                module = subj.modules[module_no]
                for phrase in module.topics:
                    total += 1
                    not_covered += 1
                    row = {
                        "semester": subj.semester,
                        "code": code,
                        "subject_name": subj.display_title,
                        "module_no": module_no,
                        "module_title": module.module_title,
                        "syllabus_topic": phrase,
                        "matched_topic": "",
                        "score": 0.0,
                        "status": "Not Covered",
                        "topic_path": "",
                        "rating": "",
                    }
                    detailed_topic_rows.append(row)
            ws_missing.append(
                [
                    subj.semester,
                    code,
                    subj.display_title,
                    subj.exam_type,
                    notes,
                ]
            )
        else:
            subject_name = app_subj.name or subj.display_title
            for module_no in sorted(subj.modules.keys()):
                module = subj.modules[module_no]
                module_candidates = app_subj.modules.get(module_no, {}).get("topics", [])
                if not isinstance(module_candidates, list):
                    module_candidates = []

                # fallback to all subject topics
                if not module_candidates:
                    module_candidates = app_subj.topics

                for phrase in module.topics:
                    total += 1
                    best_topic, score = best_match_for_phrase(phrase, module_candidates)
                    status = coverage_status(score)
                    if status == "Covered":
                        covered += 1
                    elif status == "Partial":
                        partial += 1
                    else:
                        not_covered += 1

                    if best_topic is not None:
                        ratings.append(best_topic.rating)
                        matched_title = best_topic.topic_title
                        matched_path = best_topic.topic_path
                        matched_rating = best_topic.rating
                    else:
                        matched_title = ""
                        matched_path = ""
                        matched_rating = ""

                    row = {
                        "semester": app_subj.semester,
                        "code": code,
                        "subject_name": subject_name,
                        "module_no": module_no,
                        "module_title": module.module_title,
                        "syllabus_topic": phrase,
                        "matched_topic": matched_title,
                        "score": score,
                        "status": status,
                        "topic_path": matched_path,
                        "rating": matched_rating,
                    }
                    detailed_topic_rows.append(row)

        coverage_pct = 0.0
        if total > 0:
            coverage_pct = round(((covered + (0.5 * partial)) / total) * 100, 2)

        fully_covered = "Yes" if total > 0 and not_covered == 0 else "No"

        avg_rating = ""
        if app_subj is not None and app_subj.topics:
            avg_rating = round(statistics.mean([t.rating for t in app_subj.topics]), 2)

        ws_sum.append(
            [
                subj.semester or (app_subj.semester if app_subj else ""),
                code,
                (app_subj.name if app_subj else subj.display_title),
                "Yes" if app_subj else "No",
                total,
                covered,
                partial,
                not_covered,
                coverage_pct,
                fully_covered,
                avg_rating,
                notes,
            ]
        )

    # Add app subjects that were not detected in syllabus parse as explicit note.
    for code in sorted(app_subjects.keys(), key=lambda c: (app_subjects[c].semester, c)):
        if code in theory_syllabus_codes:
            continue
        app_subj = app_subjects[code]
        ws_sum.append(
            [
                app_subj.semester,
                code,
                app_subj.name,
                "Yes",
                0,
                0,
                0,
                0,
                "",
                "No",
                round(statistics.mean([t.rating for t in app_subj.topics]), 2) if app_subj.topics else "",
                "No matching theory syllabus section found in provided PDFs",
            ]
        )

    for row in detailed_topic_rows:
        ws_topic.append(
            [
                row["semester"],
                row["code"],
                row["subject_name"],
                row["module_no"],
                row["module_title"],
                row["syllabus_topic"],
                row["matched_topic"],
                row["score"],
                row["status"],
                row["topic_path"],
                row["rating"],
                "",
            ]
        )

    for row_num in range(2, ws_topic.max_row + 1):
        status = str(ws_topic.cell(row=row_num, column=9).value)
        if status == "Covered":
            fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        elif status == "Partial":
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        else:
            fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        ws_topic.cell(row=row_num, column=9).fill = fill

    for row_num in range(2, ws_sum.max_row + 1):
        fully = str(ws_sum.cell(row=row_num, column=10).value)
        if fully == "Yes":
            fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        else:
            fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        ws_sum.cell(row=row_num, column=10).fill = fill

    autosize_columns(ws_sum)
    autosize_columns(ws_topic)
    autosize_columns(ws_missing)

    wb_cov.save(coverage_path)

    return review_path, coverage_path, subject_summary, detailed_topic_rows


def main() -> None:
    # Parse syllabus text from all provided PDFs.
    all_syllabus_subjects: List[SyllabusSubject] = []
    for key, pdf in SYLLABUS_PDFS.items():
        text = extract_pdf_text(pdf, key)
        parsed = parse_syllabus_subjects(text, key)
        all_syllabus_subjects.extend(parsed)

    syllabus_by_code = build_syllabus_maps(all_syllabus_subjects)
    app_subjects = load_app_subjects()

    review_path, coverage_path, subject_summary, detailed_rows = generate_workbooks(
        app_subjects,
        syllabus_by_code,
    )

    print("Generated files:")
    print(f"- {review_path}")
    print(f"- {coverage_path}")
    print()
    print(f"App subjects parsed: {len(app_subjects)}")
    print(f"Syllabus subjects mapped by code: {len(syllabus_by_code)}")
    print(f"Detailed syllabus topic rows: {len(detailed_rows)}")

    covered_subjects = 0
    for code, summary in subject_summary.items():
        if summary["not_covered_topics"] == 0:
            covered_subjects += 1
    print(f"Subjects with no 'Not Covered' app-topic tag (heuristic): {covered_subjects}/{len(subject_summary)}")


if __name__ == "__main__":
    main()
